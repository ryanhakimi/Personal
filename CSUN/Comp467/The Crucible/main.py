import csv
import argparse
from typing import List, Dict, Tuple
from pymongo import MongoClient
import subprocess
import json
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
import os
import requests

## function definitions ##

# normalize path
def strip_storage_prefix(path: str) -> str:

    parts = path.strip().split("/")

    # remove empty segments
    parts = [p for p in parts if p]

    if not parts:
        return ""

    if "production" in parts:
        idx = parts.index("production") + 1
    else:
        idx = 1 if len(parts) > 1 else 0

    norm = "/".join(parts[idx:])
    return norm


# baselight parsing
def load_baselight_export(line: str) -> Dict:

    line = line.strip()
    if not line:
        return None

    tokens = line.split()
    path = tokens[0]
    frame_strs = tokens[1:]

    frames: List[int] = []
    for tok in frame_strs:
        try:
            frames.append(int(tok))
        except ValueError:
            continue

    return {
        "raw_path": path,
        "norm_path": strip_storage_prefix(path),
        "frames": frames,
    }


def parse_baselight_file(path: str) -> List[Dict]:

    entries: List[Dict] = []
    with open(path, "r") as f:
        for line in f:
            entry = load_baselight_export(line)
            if entry:
                entries.append(entry)
    return entries


# xytech parsing
def load_xytech_locations(path: str) -> List[Dict]:

    entries: List[Dict] = []
    with open(path, "r") as f:
        in_locations = False
        for raw_line in f:
            line = raw_line.rstrip("\n")

            if not in_locations:
                if line.strip().startswith("Location:"):
                    in_locations = True
                continue

            # inside location section
            if not line.strip():
                break

            location_path = line.strip()
            entries.append({
                "raw_path": location_path,
                "norm_path": strip_storage_prefix(location_path),
            })

    return entries


# frame to range helpers
def frames_to_ranges(frames: List[int]) -> List[Tuple[int, int]]:

    if not frames:
        return []

    frames = sorted(set(frames))
    ranges: List[Tuple[int, int]] = []

    start = prev = frames[0]
    for num in frames[1:]:
        if num == prev + 1:
            prev = num
        else:
            ranges.append((start, prev))
            start = prev = num
    ranges.append((start, prev))

    return ranges


def format_range(r: Tuple[int, int]) -> str:

    a, b = r
    return f"{a}" if a == b else f"{a}-{b}"


# matching and csv export
def build_match_table(
    baselight_entries: List[Dict],
    xytech_entries: List[Dict],
) -> List[Dict]:

    # build a lookup from stripped path to matching entries
    bl_map: Dict[str, List[Dict]] = {}
    for e in baselight_entries:
        bl_map.setdefault(e["norm_path"], []).append(e)

    matches: List[Dict] = []
    for x in xytech_entries:
        norm = x["norm_path"]
        if norm in bl_map:
            frames: List[int] = []
            for e in bl_map[norm]:
                frames.extend(e["frames"])

            ranges = frames_to_ranges(frames)
            formatted_ranges = [format_range(r) for r in ranges]

            matches.append({
                "xytech_path": x["raw_path"],
                "norm_path": norm,
                "frame_ranges": formatted_ranges,
            })

    return matches


def write_matches_to_csv(matches: List[Dict], output_csv_path: str) -> None:

    with open(output_csv_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Location", "FrameRange"])

        for m in matches:
            for fr in m["frame_ranges"]:
                writer.writerow([m["xytech_path"], fr])


# mongo helpers
def get_db():
    client = MongoClient("mongodb://localhost:27017/")
    return client["proj4_db"]


def save_baselight_to_db(db, baselight_entries, source_name):
    coll = db["baselight"]

    docs = []
    for e in baselight_entries:
        docs.append({
            "source": source_name,
            "raw_path": e["raw_path"],
            "norm_path": e["norm_path"],
            "frames": e["frames"],
        })

    if docs:
        coll.insert_many(docs)


def save_xytech_to_db(db, xytech_entries, source_name):
    coll = db["xytech"]

    docs = []
    for x in xytech_entries:
        docs.append({
            "source": source_name,
            "raw_path": x["raw_path"],
            "norm_path": x["norm_path"],
        })

    if docs:
        coll.insert_many(docs)


def process_video(video_path: str, vimeo_token=None):

    print(f"yo, gonna process video: {video_path}")

    db = get_db()

    ps_entries = get_planeshifter_entries(db)
    print(f"found {len(ps_entries)} planeshifter entries")

    all_frames = []
    for e in ps_entries:
        all_frames.extend(e["frames"])
    
    # only print unique frames
    all_frames = sorted(set(all_frames))

    ps_ranges = add_handles_to_frames(all_frames)

    print("ranges w/ handles:")
    for r in ps_ranges:
        print(r)
    
    # pull fps + starting tc from video
    fps, start_tc = get_video_info(video_path)
    print(f"video fps looks like: {fps}")
    if start_tc:
        print(f"video start timecode: {start_tc}")
        base_frame = timecode_to_frames(start_tc, fps)
    else:
        print("no timecode found, assuming base 00:00:00:00")
        base_frame = 0

    # convert frame ranges to tc ranges
    tc_ranges = []
    for idx, (start_f, end_f) in enumerate(ps_ranges, start=1):
        start_tc_str = frames_to_timecode(start_f, fps, base_frame)
        end_tc_str = frames_to_timecode(end_f, fps, base_frame)

        mid_f = (start_f + end_f) // 2
        thumb_path = make_thumbnail(video_path, mid_f, fps, idx)  # you already have this

        clip_path = render_clip(video_path, start_f, end_f, fps, idx)

        upload_info = None
        if vimeo_token:
            # give each clip a simple title
            clip_title = f"planeshifter_{idx:02d}"
            try:
                upload_info = upload_clip_to_vimeo(vimeo_token, clip_path, clip_title)
            except Exception as e:
                print(f"vimeo upload for {clip_path} failed: {e}")

        tc_ranges.append({
            "start_frame": start_f,
            "end_frame": end_f,
            "start_tc": start_tc_str,
            "end_tc": end_tc_str,
            "thumb_path": thumb_path,
            "clip_path": clip_path,
            "vimeo": upload_info,
        })

    print("ranges as timecode:")
    for r in tc_ranges:
        print(
            f"{r['start_frame']}–{r['end_frame']} "
            f"-> {r['start_tc']} to {r['end_tc']}"
        )
    
    return tc_ranges


def get_planeshifter_entries(db):
    coll = db["baselight"]
    # searching by substring in norm_path
    return list(coll.find({"norm_path": {"$regex": "Planeshifter"}}))


def add_handles_to_frames(frames, fps=24, seconds=2):
    
    if not frames:
        return []

    frames = sorted(frames)

    handle = fps * seconds
    ranges = []

    for f in frames:
        start = f - handle
        end = f + handle
        if start < 0:
            start = 0
        ranges.append((start, end))

    return ranges


def get_video_info(video_path: str):
    
    cmd = [
        "ffprobe",
        "-v", "error",
        "-select_streams", "v:0",
        "-show_entries", "stream=avg_frame_rate:format_tags=timecode",
        "-of", "json",
        video_path,
    ]

    result = subprocess.run(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )

    if result.returncode != 0:
        print("ffprobe had an issue, falling back to 24fps, no tc")
        return 24.0, None

    data = json.loads(result.stdout or "{}")

    fps = 24.0
    timecode_str = None

    streams = data.get("streams", [])
    if streams:
        s0 = streams[0]
        fr = s0.get("avg_frame_rate")
        if fr and fr != "0/0":
            num, den = fr.split("/")
            try:
                fps = float(num) / float(den)
            except ZeroDivisionError:
                fps = 24.0

    # timecode might sit under format.tags.timecode
    fmt = data.get("format", {})
    tags = fmt.get("tags", {})
    tc = tags.get("timecode")
    if tc:
        timecode_str = tc

    return fps, timecode_str


def make_thumbnail(video_path, frame, fps, idx):
    # pick a time in seconds for ffmpeg
    seconds = frame / fps if fps > 0 else frame / 24.0

    thumb_name = f"thumb_{idx:02d}.png"

    cmd = [
        "ffmpeg",
        "-y",                # overwrite
        "-ss", str(seconds),
        "-i", video_path,
        "-vframes", "1",
        "-s", "96x74",
        thumb_name,
    ]

    # keep output quiet
    subprocess.run(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )

    return thumb_name


def timecode_to_frames(tc: str, fps: float) -> int:
    
    parts = tc.split(":")
    if len(parts) != 4:
        return 0

    h = int(parts[0])
    m = int(parts[1])
    s = int(parts[2])
    f = int(parts[3])

    total = (((h * 60) + m) * 60 + s) * fps + f
    return int(round(total))


def frames_to_timecode(frame: int, fps: float, base_frame: int = 0) -> str:
    
    total = frame + base_frame
    fps_i = int(round(fps))
    if fps_i <= 0:
        fps_i = 24

    total = int(round(total))

    secs, ff = divmod(total, fps_i)
    mins, ss = divmod(secs, 60)
    hrs, mm = divmod(mins, 60)

    return f"{hrs:02d}:{mm:02d}:{ss:02d}:{ff:02d}"


def write_xls_with_planeshifter(output_xls_path, matches, tc_ranges):
    wb = Workbook()
    ws = wb.active
    ws.title = "stuff"

    # main match table, same as csv
    ws.append(["Location", "FrameRange"])
    for m in matches:
        loc = m["xytech_path"]
        for fr in m["frame_ranges"]:
            ws.append([loc, fr])

    # add planeshifter tc stuff as extra section
    if tc_ranges:
        ws.append([])  # blank row
        ws.append(["ps_start_frame", "ps_end_frame", "ps_start_tc", "ps_end_tc", "ps_thumb"])

        for r in tc_ranges:
            ws.append([
                r["start_frame"],
                r["end_frame"],
                r["start_tc"],
                r["end_tc"],
                "",  # placeholder for thumb
            ])

        # drop thumbnails into the last column
        start_row = ws.max_row - len(tc_ranges) + 1
        thumb_col = 5  # column E

        for i, r in enumerate(tc_ranges):
            thumb_path = r.get("thumb_path")
            if not thumb_path:
                continue

            row = start_row + i
            cell_ref = ws.cell(row=row, column=thumb_col).coordinate

            try:
                img = XLImage(thumb_path)
                # sizes are already 96x74, but confirm
                img.width = 96
                img.height = 74
                ws.add_image(img, cell_ref)
            except Exception as e:
                print(f"couldnt add img {thumb_path}: {e}")

    wb.save(output_xls_path)
    print(f"wrote xls to {output_xls_path}")


def render_clip(video_path, start_frame, end_frame, fps, idx):
    
    # figure out time in seconds
    if fps <= 0:
        fps = 24.0

    start_sec = start_frame / fps
    duration = (end_frame - start_frame) / fps

    out_name = f"ps_clip_{idx:02d}.mp4"

    cmd = [
        "ffmpeg",
        "-y",               # overwrite
        "-ss", str(start_sec),
        "-i", video_path,
        "-t", str(duration),
        "-c:v", "libx264",
        "-c:a", "aac",
        out_name,
    ]

    subprocess.run(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )

    return out_name


def vimeo_headers(token):
    return {
        "Authorization": f"bearer {token}",
        "Accept": "application/vnd.vimeo.*+json;version=3.4",
    }


def upload_clip_to_vimeo(token, file_path, title):
    size = os.path.getsize(file_path)

    # create video object with tus upload
    create_data = {
        "upload": {
            "approach": "tus",
            "size": str(size),
        },
        "name": title,
    }

    resp = requests.post(
        "https://api.vimeo.com/me/videos",
        headers={**vimeo_headers(token), "Content-Type": "application/json"},
        json=create_data,
    )
    resp.raise_for_status()
    info = resp.json()

    upload = info.get("upload", {})
    upload_link = upload.get("upload_link")
    uri = info.get("uri")
    link = info.get("link")

    if not upload_link:
        print("no upload_link from vimeo, upload failed?")
        return None

    # send the file with tus
    with open(file_path, "rb") as f:
        data = f.read()

    tus_headers = {
        "Tus-Resumable": "1.0.0",
        "Upload-Offset": "0",
        "Content-Type": "application/offset+octet-stream",
        "Accept": "application/vnd.vimeo.*+json;version=3.4",
    }

    tus_resp = requests.patch(upload_link, headers=tus_headers, data=data)
    tus_resp.raise_for_status()

    print(f"uploaded {file_path} to vimeo: {uri} ({link})")
    return {
        "title": title,
        "uri": uri,
        "link": link,
    }


def write_vimeo_csv(token, csv_path):
    url = "https://api.vimeo.com/me/videos"
    headers = vimeo_headers(token)

    resp = requests.get(url, headers=headers, params={"per_page": 50})
    resp.raise_for_status()
    data = resp.json()

    items = data.get("data", [])

    with open(csv_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Title", "URI", "PublicLink", "Status"])

        for vid in items:
            title = vid.get("name")
            uri = vid.get("uri")
            link = vid.get("link")
            # status can live under transcode or upload depending on account/settings
            transcode = vid.get("transcode", {}) or {}
            upload = vid.get("upload", {}) or {}
            status = transcode.get("status") or upload.get("status")

            w.writerow([title, uri, link, status])

    print(f"wrote vimeo csv to {csv_path}")

## func def end ##


parser = argparse.ArgumentParser(
    description="match baselight export to xytech locations and export csv"
)

parser.add_argument(
    "--baselight",
    required=True,
    help="path to baselight export text file"
)

parser.add_argument(
    "--xytech",
    required=True,
    help="path to xytech workorder text file"
)

parser.add_argument(
    "--process",
    help="video file to process (trailer demo)"
)

parser.add_argument(
    "--output",
    help="excel file (xlsx) to write"
)

parser.add_argument(
    "--vimeo_token",
    help="vimeo personal access token for upload/api"
)

parser.add_argument(
    "--vimeo_csv",
    help="where to dump vimeo video info as csv"
)

args = parser.parse_args()

baselight_entries = parse_baselight_file(args.baselight)
xytech_entries = load_xytech_locations(args.xytech)

matches = build_match_table(baselight_entries, xytech_entries)

output_csv = "match_output.csv"
write_matches_to_csv(matches, output_csv)

# stash in mongo
db = get_db()
save_baselight_to_db(db, baselight_entries, args.baselight)
save_xytech_to_db(db, xytech_entries, args.xytech)

tc_ranges = None
if args.process:
    tc_ranges = process_video(args.process, vimeo_token=args.vimeo_token)

if args.output:
    write_xls_with_planeshifter(args.output, matches, tc_ranges)

if args.vimeo_token and args.vimeo_csv:
    write_vimeo_csv(args.vimeo_token, args.vimeo_csv)